
from openpyxl import Workbook #for writing in spreadsheet

from openpyxl import load_workbook #for reading from spreadsheet


techcrunch = load_workbook(filename='Data.xlsx')

for sheets in techcrunch.sheetnames:
    data_sheet = techcrunch.active(sheets)



